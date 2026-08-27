"""FEC "Federal Elections" — who actually won (PRD FR-C2, `election_result`).

openFEC is a campaign-finance disclosure API and carries no outcome at all, so
`campaign_finance.election_result` has to come from somewhere else. The P4
design note (§8-A) chose OpenElections `fec_results` for it. Measured
2026-08-27, that choice does not survive contact:

  * `openelections/fec_results` is a RUBY LIBRARY, and its static JSON API
    (openelections.github.io/fec_results) publishes 2000, 2002 ... 2014 and
    stops. The five-year window this milestone covers — 2022, 2024, 2026 — is
    entirely outside it.
  * What that library actually does is download and parse the FEC's own
    `federalelections{year}.xlsx` workbook, sheet "US House Results by State".
    Its README says so and its `process_2014` reads exactly that sheet.

So the substitution here is not a different source, it is the SAME source
without the wrapper: the FEC's own publication, one hop closer, still a
first-party official record (PRD FC-1). That also removes the name-matching
step the design note was worried about — the FEC workbook carries `FEC ID`, so
the join to `candidate.fec_candidate_id` is exact and a fuzzy fallback is never
needed for results.

WHAT IS AND IS NOT PUBLISHED
----------------------------
Measured 2026-08-27 against fec.gov's own election-results index:

| election | what the FEC publishes                        | what we can say      |
|----------|-----------------------------------------------|----------------------|
| 2022     | federalelections2022.xlsx — full results      | W / L / N            |
| 2024     | 2024congressgecands.xlsx — general BALLOT     | N only (see below)   |
| 2026     | nothing; the election is in November 2026     | nothing              |

The "Federal Elections" compilation series runs ...2018, 2022 — there is no
`federal-elections-2024` page (404), and no 2024 results workbook. What exists
for 2024 is the general-election ballot list: every candidate who appeared on
the November ballot, with an FEC ID, and no votes and no winner.

That list still answers exactly one of the three outcomes honestly. A
candidate who filed for 2024 and is ABSENT from it did not reach the general
ballot, which is 'N'. A candidate who is present either won or lost, and the
FEC has not said which — so their result stays NULL. The UI must therefore
read a NULL by cycle, not by row (FR-C4): in 2022 a NULL means the FEC's
compilation had no row for that candidate, in 2024 it means the outcome is not
published yet, in 2026 it means the election has not happened.

MEASURED SHAPE OF THE 2022 WORKBOOK
-----------------------------------
`GE WINNER INDICATOR` = 'W' marks the winner: 484 rows on the House sheet, 40
on the Senate sheet. Three details the naive read gets wrong:

1. `GENERAL VOTES` is empty for 22 House and 2 Senate winners. Those are
   fusion-ballot candidates whose votes are split across party lines and
   totalled in `COMBINED GE PARTY TOTALS`; two more carry the literal string
   'Unopposed'. So "was on the general ballot" is not "has a vote count".
2. A candidate can appear TWICE for one year. California ran a full-term and
   an unexpired-term Senate election on the same 2022 ballot, and Alex Padilla
   won both; one House candidate appears in a special and the regular election.
   Outcomes are merged best-first (W beats L beats N), never last-write-wins.
3. Rows whose `FEC ID` is 'n/a' are district subtotals and write-in
   aggregates ("Scattered", "District Votes:"), not candidates.

Whitespace: IDs in these sheets carry trailing spaces and, in the 2024 NC
sheet, a leading non-breaking space. They are stripped before joining — an ID
that differs by whitespace joins to nothing and looks like missing data.
"""

from __future__ import annotations

import io
from collections.abc import Collection, Iterator
from dataclasses import dataclass
from typing import Any

from common.http import Fetcher, build_client
from common.logging import get_logger
from sources.base import FetchResult, SourceError

log = get_logger(__name__)

FEC_BASE_URL = "https://www.fec.gov"

# Explicit, verified document URLs — never a guessed one. The FEC serves these
# from an opaque numeric document id, so a new year cannot be extrapolated: a
# guessed id is either a 404 or, worse, some other year's file silently loaded
# as this year's results. The pattern follows census_tiger's vintage table.
#
# Index: https://www.fec.gov/introduction-campaign-finance/election-results-and-voting-information/
RESULTS_URL_BY_YEAR: dict[int, str] = {
    2022: f"{FEC_BASE_URL}/documents/5676/federalelections2022.xlsx",
}

# The general-election ballot list, published ahead of the results compilation.
BALLOT_URL_BY_YEAR: dict[int, str] = {
    2024: f"{FEC_BASE_URL}/documents/5548/2024congressgecands.xlsx",
}

# Sheet names are not stable and the year is not always in the same place:
# '8. US House Results by State' (2022), '2014 US House Results by State'
# (2014), 'WY General Ballot 2024', 'NV General BallotS 2024',
# 'CA SEN Spec Gen Ballot 2024', 'WI 08 Special General Ballot'. So a sheet is
# found by a substring of its name, never by an exact name and never by a
# suffix — the ballot workbook puts the year last on 56 of its 58 sheets, and
# a suffix match silently read only the two that did not.
HOUSE_SHEET_MATCH = "us house results by state"
SENATE_SHEET_MATCH = "us senate results by state"
BALLOT_SHEET_MATCH = "ballot"

# `election_result` values, best first. A candidate with two rows for one year
# keeps the best of them.
RESULT_RANK = {"W": 0, "L": 1, "N": 2}


def open_fetcher() -> Fetcher:
    """A fetcher for www.fec.gov. No key; the workbooks are a few MB."""
    return Fetcher(build_client(timeout=180.0), source_name="fec_results")


def results_url(year: int) -> str:
    """URL of the Federal Elections workbook for one election year."""
    try:
        return RESULTS_URL_BY_YEAR[year]
    except KeyError:
        known = ", ".join(str(y) for y in sorted(RESULTS_URL_BY_YEAR))
        raise SourceError(
            f"the FEC has not published a Federal Elections compilation for {year} "
            f"that this pipeline has verified (known: {known}). Check "
            f"{FEC_BASE_URL}/introduction-campaign-finance/"
            f"election-results-and-voting-information/ and add the document URL "
            f"to RESULTS_URL_BY_YEAR."
        ) from None


def fetch_results(fetcher: Fetcher, *, year: int) -> FetchResult:
    """Download the Federal Elections workbook for one election year."""
    url = results_url(year)
    log.info("fec_results.fetch", url=url, year=year)
    result = fetcher.get(url)
    _assert_xlsx(result, url)
    return result


def fetch_ballot(fetcher: Fetcher, *, year: int) -> FetchResult:
    """Download the general-election ballot list for one election year."""
    try:
        url = BALLOT_URL_BY_YEAR[year]
    except KeyError:
        known = ", ".join(str(y) for y in sorted(BALLOT_URL_BY_YEAR))
        raise SourceError(
            f"no verified general-election ballot list for {year} (known: {known})"
        ) from None
    log.info("fec_results.fetch_ballot", url=url, year=year)
    result = fetcher.get(url)
    _assert_xlsx(result, url)
    return result


def _assert_xlsx(result: FetchResult, url: str) -> None:
    # .xlsx is a zip. A 200 that is really an HTML error page would otherwise
    # reach openpyxl and fail as a confusing BadZipFile much later.
    if not result.payload.startswith(b"PK"):
        raise SourceError(f"{url} did not return an .xlsx workbook")


@dataclass(frozen=True, slots=True)
class ElectionOutcome:
    """One candidate's outcome in one election, as the FEC published it."""

    fec_candidate_id: str
    election_year: int
    office: str
    state: str
    # The district as printed. Numeric for the House ('00'..'53'); for the
    # Senate the FEC prints 'S', 'S-Full Term' or 'S-Unexpired Term', which is
    # a seat label and not a district.
    district_label: str
    result: str
    on_general_ballot: bool


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).replace("\xa0", " ").split())
    return cleaned or None


def _candidate_id(value: Any) -> str | None:
    """An FEC ID from a spreadsheet cell, or None for the non-candidate rows."""
    cleaned = _text(value)
    if cleaned is None:
        return None
    cleaned = cleaned.replace(" ", "").upper()
    # Subtotal and write-in rows carry 'n/a'.
    return None if cleaned in ("N/A", "NA", "") else cleaned


def _sheet_rows(payload: bytes, *, match: str) -> Iterator[tuple[list[str], tuple[Any, ...]]]:
    """Yield `(header, row)` for every sheet whose name CONTAINS `match`.

    Read-only streaming: the 2022 workbook is 3 MB compressed and the House
    sheet alone is 4,471 rows wide of 30 columns.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        names = [n for n in workbook.sheetnames if match in n.strip().lower()]
        if not names:
            raise SourceError(
                f"workbook has no sheet named like {match!r} — got {workbook.sheetnames}"
            )
        for name in names:
            rows = workbook[name].iter_rows(values_only=True)
            try:
                header = [(_text(h) or "").upper() for h in next(rows)]
            except StopIteration:
                continue
            for row in rows:
                yield header, row
    finally:
        workbook.close()


def _column(header: list[str], *names: str) -> int | None:
    for name in names:
        if name in header:
            return header.index(name)
    return None


def parse_results(
    payload: bytes,
    *,
    year: int,
    states: Collection[str] | None = None,
) -> Iterator[ElectionOutcome]:
    """Yield one `ElectionOutcome` per candidate row of a results workbook.

    Args:
        payload: the raw .xlsx bytes, exactly as fetched.
        year: stamped onto every outcome; the workbook itself is per-year.
        states: two-letter codes to keep. None keeps every state.

    Both chamber sheets are read. Rows without an FEC ID are subtotals, not
    candidates, and are skipped.
    """
    wanted = {s.upper() for s in states} if states is not None else None

    for office, sheet_match in (("H", HOUSE_SHEET_MATCH), ("S", SENATE_SHEET_MATCH)):
        for header, row in _sheet_rows(payload, match=sheet_match):
            id_col = _column(header, "FEC ID", "FEC ID#")
            state_col = _column(header, "STATE ABBREVIATION")
            district_col = _column(header, "DISTRICT")
            winner_col = _column(header, "GE WINNER INDICATOR")
            votes_col = _column(header, "GENERAL VOTES")
            combined_col = _column(header, "COMBINED GE PARTY TOTALS (WHEN APPLICABLE)")
            if id_col is None or state_col is None or winner_col is None:
                raise SourceError(
                    f"{year} {office} results sheet is missing an expected column; header={header}"
                )

            candidate_id = _candidate_id(_cell(row, id_col))
            state = _text(_cell(row, state_col))
            if candidate_id is None or state is None:
                continue
            state = state.upper()
            if wanted is not None and state not in wanted:
                continue

            won = (_text(_cell(row, winner_col)) or "").upper() == "W"
            # Finding 1: a winner can have no vote count. Presence on the
            # general ballot is "the FEC printed a general-election figure of
            # any kind for this row", not "the votes column is a number".
            on_ballot = (
                won or _cell(row, votes_col) is not None or _cell(row, combined_col) is not None
            )
            yield ElectionOutcome(
                fec_candidate_id=candidate_id,
                election_year=year,
                office=office,
                state=state,
                district_label=_text(_cell(row, district_col)) or "",
                result="W" if won else ("L" if on_ballot else "N"),
                on_general_ballot=on_ballot,
            )


def parse_ballot(
    payload: bytes,
    *,
    year: int,
    states: Collection[str] | None = None,
) -> Iterator[ElectionOutcome]:
    """Yield one outcome per row of a general-election BALLOT list.

    Every row means "this candidate was on the November general-election
    ballot" and nothing more, so `result` is left empty and only
    `on_general_ballot` is set. The caller turns the ABSENCE of a candidate
    from this list into 'N'; it must never turn a presence into 'W' or 'L'.

    One sheet per state ("NC General Ballot 2024"), plus separate sheets for
    same-day special elections ("CA SEN Spec Gen Ballot 2024"), all of which
    count as being on the ballot.
    """
    wanted = {s.upper() for s in states} if states is not None else None

    for header, row in _sheet_rows(payload, match=BALLOT_SHEET_MATCH):
        id_col = _column(header, "FEC ID#", "FEC ID")
        state_col = _column(header, "STATE ABBREVIATION")
        district_col = _column(header, "DISTRICT")
        if id_col is None or state_col is None:
            raise SourceError(f"{year} ballot sheet is missing an expected column; header={header}")

        candidate_id = _candidate_id(_cell(row, id_col))
        state = _text(_cell(row, state_col))
        if candidate_id is None or state is None:
            continue
        state = state.upper()
        if wanted is not None and state not in wanted:
            continue

        district_label = _text(_cell(row, district_col)) or ""
        yield ElectionOutcome(
            fec_candidate_id=candidate_id,
            election_year=year,
            office="S" if district_label.upper().startswith("S") else "H",
            state=state,
            district_label=district_label,
            result="",
            on_general_ballot=True,
        )


def merge_outcomes(outcomes: Iterator[ElectionOutcome]) -> dict[str, str]:
    """Collapse outcomes to one result per candidate.

    A workbook covers one election year, so the candidate id is the whole key.

    Finding 2: a candidate can hold two rows for that year — California's 2022
    full-term and unexpired-term Senate elections, or a special held alongside
    the regular one. The best outcome wins, so a candidate who won one and lost
    the other is recorded as having won, never as having lost.
    """
    best: dict[str, str] = {}
    for outcome in outcomes:
        if not outcome.result:
            continue
        current = best.get(outcome.fec_candidate_id)
        if current is None or RESULT_RANK[outcome.result] < RESULT_RANK[current]:
            best[outcome.fec_candidate_id] = outcome.result
    return best
